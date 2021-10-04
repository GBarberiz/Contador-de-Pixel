from PIL import Image
import csv
import os
import os.path
from openpyxl import Workbook
#from openpyxl import Workbook

imName = ""
arr = os.listdir()
for file in arr:
    extension = os.path.splitext(file)[1]
    if extension == ".png":
        imName = os.path.splitext(file)[0] + os.path.splitext(file)[1]

imFile = Image.open(imName)

pixelColor = {}
for pixel in imFile.getdata():
    if pixelColor.get(pixel):
        pixelColor[pixel]+=1

    else:
        pixelColor[pixel]=1

#csv
field_names = ['Color',"Frequency"]
with open(imName+".csv", 'w') as csv_file: 
    writer = csv.DictWriter(csv_file, fieldnames = field_names) 
    writer.writeheader() 
    for pixelRGB, pixelFrequency in pixelColor.items():
        writer.writerow({"Color": pixelRGB,"Frequency": pixelFrequency})

#xlsx
wb = Workbook()
ws =  wb.active
ws.title = "Colors"

cellref=ws.cell(row=1, column=1, value = "Color")
cellref=ws.cell(row=1, column=3, value = "Frequency")

r = 2
c = 1
for pixelRGB, pixelFrequency in pixelColor.items():

    cellref=ws.cell(row=r, column=c,value = str(pixelRGB))
    c+=2
    cellref=ws.cell(row=r, column=c,value = pixelFrequency)
    r+=1
    c-=2

wb.save(filename = imName+".xlsx")
        
